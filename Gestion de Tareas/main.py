from datetime import datetime
import openpyxl
from openpyxl import Workbook
import sys
from tkinter import Tk, filedialog

tareas = []  # Lista de tareas, inicialmente vacía
contador_tareas = 1  # Puntero


def mostrar_menu():
    print("1. Agregar una Nueva Tarea")
    print("2. Marcar una Tarea como Completada")
    print("3. Ver Tareas Pendientes")
    print("4. Ver Tareas Completadas")
    print("5. Ver Todas las Tareas")
    print("6. Eliminar una Tarea")
    print("7. Guardar Tareas")
    print("8. Cargar Tareas")
    print("9. Salir del programa")


def agregar_tarea():
    global contador_tareas
    nombre = input("Ingrese el nombre de la tarea: ")
    descripcion = input("Ingrese una descripción opcional de la tarea (puede dejarlo en blanco): ")
    identificador = contador_tareas
    fecha_creacion = datetime.now()
    estado = "Pendiente"

    tarea = {
        "id": identificador,
        "nombre": nombre,
        "descripcion": descripcion,
        "fecha_creacion": fecha_creacion,
        "estado": estado
    }

    tareas.append(tarea)
    contador_tareas += 1
    print("Tarea agregada con éxito")


def marcar_completada():
    mostrar_tareas_pendientes()
    try:
        opcion = int(input("Seleccione el número de la tarea a marcar como completada: "))
        tarea = tareas[opcion - 1]
        tarea['estado'] = "Completado"
        print(f"La tarea '{tarea['nombre']}' ha sido marcada como completada.")

    except (ValueError, IndexError):
        print("¡Opción inválida! Por favor, seleccione un número válido.")


def mostrar_tareas_pendientes():
    print("Tareas pendientes:")
    for i, tarea in enumerate(tareas):
        if tarea.get('estado') == 'Pendiente':
            print(f"{i + 1}. Nombre: {tarea['nombre']}, Descripción: {tarea['descripcion']}, Estado: {tarea['estado']}")


def mostrar_tareas_completadas():
    print("Tareas completadas:")
    for i, tarea in enumerate(tareas):
        if tarea.get('estado') == 'Completado':
            print(f"{i + 1}. Nombre: {tarea['nombre']}, Descripción: {tarea['descripcion']}, Estado: {tarea['estado']}")


def mostrar_tareas():
    print("Todas las tareas:")
    for i, tarea in enumerate(tareas):
        print(f"{i + 1}. Nombre: {tarea['nombre']}, Descripción: {tarea['descripcion']}, Estado: {tarea['estado']}")


def eliminar_tarea():
    mostrar_tareas()
    try:
        opcion = int(input("Seleccione el número de la tarea a eliminar: "))
        tarea = tareas.pop(opcion - 1)
        print(f"La tarea '{tarea['nombre']}' ha sido eliminada.")
    except (ValueError, IndexError):
        print("¡Opción inválida! Por favor, seleccione un número válido.")


def guardar_tareas():
    root = Tk()
    root.withdraw()  # Ocultar la ventana principal

    # Mostrar el diálogo para seleccionar la ubicación y nombre del archivo
    nombre_archivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos de Excel", "*.xlsx")])

    if not nombre_archivo:
        # Si el usuario cancela la operación, salir de la función
        return

    libro = Workbook()
    hoja = libro.active
    hoja.append(["ID", "Nombre", "Descripción", "Fecha de creación", "Estado"])

    for i, tarea in enumerate(tareas):
        fila = [i + 1, tarea.get('nombre', ''), tarea.get('descripcion', ''), tarea.get('fecha_creacion', ''),
                tarea.get('estado', '')]
        hoja.append(fila)

    libro.save(nombre_archivo)


def cargar_tareas():
    try:
        # Crear una instancia de Tkinter
        root = Tk()
        root.withdraw()  # Ocultar la ventana principal

        # Abrir el cuadro de diálogo para seleccionar el archivo
        filepath = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel files", "*.xlsx")]
        )

        if not filepath:  # Si el usuario cancela la selección
            return []

        # Cargar el libro de trabajo (workbook)
        libro = openpyxl.load_workbook(filepath)
        hoja = libro.active

        for fila in hoja.iter_rows(min_row=2, values_only=True):
            tarea = {
                "nombre": fila[1],
                "descripcion": fila[2],
                "fecha_creacion": fila[3],
                "estado": fila[4]
            }
            tareas.append(tarea)

        print("Tareas cargadas correctamente desde el archivo.")
    except FileNotFoundError:
        print("No se encontró el archivo.")


def salir_del_programa():
    print("¡Hasta luego!")
    sys.exit(0)


print(20 * '-', 'Bienvenido Gestion Calendar', 20 * '-')
while True:
    mostrar_menu()
    print(50 * '-')
    seleccion = input("Por favor, elige una opción: ")
    print(50 * '-')
    if seleccion == "1":
        agregar_tarea()
    elif seleccion == "2":
        marcar_completada()
    elif seleccion == "3":
        mostrar_tareas_pendientes()
    elif seleccion == "4":
        mostrar_tareas_completadas()
    elif seleccion == "5":
        mostrar_tareas()
    elif seleccion == "6":
        eliminar_tarea()
    elif seleccion == "7":
        guardar_tareas()
    elif seleccion == "8":
        cargar_tareas()
    elif seleccion == "9":
        salir_del_programa()
        print("Saliendo del programa...")
        break
    else:
        print("Opción no válida. Por favor, elige una opción válida.")
